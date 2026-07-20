import re
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify
from wagtail.models import Page

from catalog.models import (
    CANONICAL_STATUS_CHOICES,
    CATEGORY_CHOICES,
    SacredSitePage,
    SaintPage,
    Topic,
)

SAINTS_PARENT_SLUG = "saints"

# Sites have no single flat parent -- each category already has its own page
# under Explore, so a site is filed under the page matching its category.
CATEGORY_PARENT_SLUGS = {
    "Marian Apparition": "marian-apparitions",
    "Eucharistic Miracle": "eucharistic-miracles",
    "Saints & Tombs": "saints-and-tombs",
    "Holy Land": "holy-lands",
    "Shrine & Basilica": "shrines-and-basilicas",
}

# The Saints tab is the General Roman Calendar (feast-day titles), so a few
# Sites-tab associated_saint values need mapping onto the calendar's exact
# title. Anything not listed here that still fails an exact match is a saint
# genuinely missing from the Saints tab -- it stays blank and gets a warning.
SAINT_ALIASES = {
    "St. Faustina Kowalska": "St. Faustina",
    "St. Therese of Lisieux": "St. Thérèse of Lisieux (Little Flower of Jesus)",
    "St. Charbel Makhlouf": "St. Sharbel Makhluf",
    "St. Mary Magdalene": "St. Mary Magdalen",
    "St. James the Greater": "St. James (Son of Zebedee)",
    "St. Paul the Apostle": "St. Paul",
    "St. Nicholas of Myra": "St. Nicholas",
    "St. Clare of Assisi": "St. Clare",
    "St. Teresa of Avila": "St. Teresa of Ávila",
    "St. Mark the Evangelist": "St. Mark",
    "St. Thomas the Apostle": "St. Thomas",
    "Sts. Isaac Jogues and Companions": "St. Isaac Jogues",
    "Pope St. John Paul II": "St. John Paul II",
}

VALID_CATEGORIES = {value for value, _ in CATEGORY_CHOICES}
VALID_CANONICAL_STATUSES = {value for value, _ in CANONICAL_STATUS_CHOICES}
TAG_RE = re.compile(r"<[a-zA-Z/][^>]*>")


def s(value):
    if value is None:
        return ""
    return str(value).strip()


def to_richtext(value):
    text = s(value)
    if not text:
        return ""
    if TAG_RE.search(text):
        return text
    return f"<p>{text}</p>"


def read_sheet(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []
    header = [s(cell) for cell in header_row]
    result = []
    for row in rows:
        if all(cell is None for cell in row):
            continue
        result.append(dict(zip(header, row)))
    return result


class Command(BaseCommand):
    help = "Bulk-import Topics, Saints, and Sites from the catalog XLSX workbook."

    def add_arguments(self, parser):
        parser.add_argument("workbook", help="Path to the .xlsx workbook.")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Run the import inside a transaction and roll it back; nothing is saved.",
        )
        parser.add_argument(
            "--only",
            default="topics,saints,sites",
            help="Comma-separated subset of sections to run, e.g. --only saints,sites",
        )

    def handle(self, *args, **options):
        workbook_path = options["workbook"]
        dry_run = options["dry_run"]
        only = {part.strip().lower() for part in options["only"].split(",") if part.strip()}
        unknown = only - {"topics", "saints", "sites"}
        if unknown:
            raise CommandError(f"Unknown --only value(s): {', '.join(sorted(unknown))}")
        try:
            workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Workbook not found: {workbook_path}")
        self.dry_run = dry_run
        self.stats = {
            "topics": {"created": 0, "updated": 0},
            "saints": {"created": 0, "updated": 0},
            "sites": {"created": 0, "updated": 0},
        }
        self.warnings = []
        self._saint_rows = []
        self._site_rows = []
        with transaction.atomic():
            if "topics" in only:
                self.import_topics(read_sheet(workbook, "Topics"))
            if "saints" in only:
                self._saint_rows = read_sheet(workbook, "Saints")
                self.import_saints(self._saint_rows)
            if "sites" in only:
                self._site_rows = read_sheet(workbook, "Sites")
                self.import_sites(self._site_rows)
            if "saints" in only and self._saint_rows:
                self.wire_saint_topics(self._saint_rows)
            if "sites" in only and self._site_rows:
                self.wire_site_relationships(self._site_rows)
            if dry_run:
                transaction.set_rollback(True)
        self.print_summary()

    def get_parent(self, slug, label):
        parent = Page.objects.filter(slug=slug).first()
        if parent is None:
            if self.dry_run:
                self.warnings.append(
                    f"Parent page with slug '{slug}' not found -- new {label} pages will be "
                    "skipped in this dry run (this will raise an error on a real run)."
                )
                return None
            raise CommandError(
                f"Parent page with slug '{slug}' not found. Create it in the page tree first, "
                "then re-run this command."
            )
        return parent.specific

    def import_topics(self, rows):
        for row in rows:
            name = s(row.get("name"))
            if not name:
                continue
            slug = s(row.get("slug")) or slugify(name)
            description = s(row.get("description"))
            topic, created = Topic.objects.update_or_create(
                slug=slug,
                defaults={"name": name, "description": description},
            )
            self.stats["topics"]["created" if created else "updated"] += 1

    def import_saints(self, rows):
        parent = self.get_parent(SAINTS_PARENT_SLUG, "saint")
        for row in rows:
            title = s(row.get("name"))
            if not title:
                continue
            slug = s(row.get("slug")) or slugify(title)
            saint = SaintPage.objects.filter(slug=slug).first()
            is_new = saint is None
            if is_new and parent is None:
                self.warnings.append(f"Skipped creating saint '{title}' -- no parent page available.")
                continue
            if is_new:
                saint = SaintPage(title=title, slug=slug)
            saint.title = title
            saint.also_known_as = s(row.get("also_known_as"))
            saint.honorific_type = s(row.get("honorific_type"))
            saint.feast_day = s(row.get("feast_day"))
            saint.born = s(row.get("born"))
            saint.died = s(row.get("died"))
            saint.canonized = s(row.get("canonized"))
            saint.patronage = s(row.get("patronage"))
            saint.significance = s(row.get("significance"))
            saint.body = to_richtext(row.get("body_draft"))
            saint.source_url = s(row.get("source_url"))
            saint.source_note = s(row.get("source_note"))
            saint.data_status = s(row.get("data_status"))
            saint.editor_notes = s(row.get("editor_notes"))
            if is_new:
                parent.add_child(instance=saint)
            else:
                saint.save()
            self.stats["saints"]["created" if is_new else "updated"] += 1

    def wire_saint_topics(self, rows):
        for row in rows:
            title = s(row.get("name"))
            if not title:
                continue
            slug = s(row.get("slug")) or slugify(title)
            saint = SaintPage.objects.filter(slug=slug).first()
            if saint is None:
                continue
            topic_names = [t.strip() for t in s(row.get("topics")).split(",") if t.strip()]
            topics = self.resolve_topics(topic_names, title)
            saint.topics.set(topics)

    def import_sites(self, rows):
        parent_cache = {}
        for row in rows:
            title = s(row.get("site_name"))
            if not title:
                continue
            slug = s(row.get("slug")) or slugify(title)

            category = s(row.get("category"))
            if category and category not in VALID_CATEGORIES:
                self.warnings.append(f"Unknown category '{category}' for site '{title}' - left blank.")
                category = ""

            parent = None
            parent_slug = CATEGORY_PARENT_SLUGS.get(category)
            if parent_slug:
                if parent_slug not in parent_cache:
                    parent_cache[parent_slug] = self.get_parent(parent_slug, f"'{category}' site")
                parent = parent_cache[parent_slug]

            site = SacredSitePage.objects.filter(slug=slug).first()
            is_new = site is None
            if is_new and parent is None:
                reason = "category is blank/unknown" if not category else f"parent page for '{category}' not found"
                self.warnings.append(f"Skipped creating site '{title}' - {reason}.")
                continue
            if is_new:
                site = SacredSitePage(title=title, slug=slug)

            canonical_status = s(row.get("canonical_status"))
            if canonical_status and canonical_status not in VALID_CANONICAL_STATUSES:
                self.warnings.append(
                    f"Unknown canonical_status '{canonical_status}' for site '{title}' - left blank."
                )
                canonical_status = ""

            site.title = title
            site.category = category
            site.canonical_status = canonical_status
            site.locality = s(row.get("locality"))
            site.country = s(row.get("country"))
            site.latitude = self.parse_decimal(row.get("latitude"), title, "latitude")
            site.longitude = self.parse_decimal(row.get("longitude"), title, "longitude")
            site.date_display = s(row.get("date_display"))
            site.feast_day = s(row.get("feast_day"))
            site.summary_short = s(row.get("summary_short"))
            site.the_story = to_richtext(row.get("the_story"))
            site.church_recognition = to_richtext(row.get("church_recognition"))
            site.catholic_teaching = to_richtext(row.get("catholic_teaching"))
            site.pilgrimage_info = to_richtext(row.get("pilgrimage_info"))
            site.go_deeper = to_richtext(row.get("go_deeper"))
            site.location_link = s(row.get("Location Link"))
            site.notes_internal = s(row.get("notes_internal"))
            # featured_image_filename intentionally ignored (filenames are notes only).

            if is_new:
                parent.add_child(instance=site)
            else:
                site.save()
            self.stats["sites"]["created" if is_new else "updated"] += 1

    def wire_site_relationships(self, rows):
        for row in rows:
            title = s(row.get("site_name"))
            if not title:
                continue
            slug = s(row.get("slug")) or slugify(title)
            site = SacredSitePage.objects.filter(slug=slug).first()
            if site is None:
                continue
            saint_name = s(row.get("associated_saint"))
            if saint_name:
                saint = SaintPage.objects.filter(title__iexact=saint_name).first()
                if saint is None and saint_name in SAINT_ALIASES:
                    saint = SaintPage.objects.filter(title__iexact=SAINT_ALIASES[saint_name]).first()
                if saint is None:
                    self.warnings.append(
                        f"Saint '{saint_name}' not found for site '{title}' -- associated_saint left blank."
                    )
                site.associated_saint = saint
            else:
                site.associated_saint = None
            topic_names = [t.strip() for t in s(row.get("topics")).split(",") if t.strip()]
            topics = self.resolve_topics(topic_names, title)
            site.topics.set(topics)
            site.save()

    def resolve_topics(self, topic_names, context_label):
        topics = []
        for name in topic_names:
            topic = Topic.objects.filter(name__iexact=name).first()
            if topic is None:
                self.warnings.append(f"Topic '{name}' not found for '{context_label}'.")
                continue
            topics.append(topic)
        return topics

    def parse_decimal(self, value, context_label, field_label):
        text = s(value)
        if not text:
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            self.warnings.append(f"Bad {field_label} value '{text}' for '{context_label}' - left blank.")
            return None

    def print_summary(self):
        if self.dry_run:
            self.stdout.write(self.style.WARNING("Dry run -- transaction rolled back, nothing was saved.\n"))
        for section in ("topics", "saints", "sites"):
            counts = self.stats[section]
            self.stdout.write(f"{section}: {counts['created']} created, {counts['updated']} updated")
        if self.warnings:
            self.stdout.write(self.style.WARNING(f"\n{len(self.warnings)} warning(s):"))
            for warning in self.warnings:
                self.stdout.write(self.style.WARNING(f"  - {warning}"))
        else:
            self.stdout.write("\n0 warnings")
